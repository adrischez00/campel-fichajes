from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime
from fastapi.responses import StreamingResponse


class ExportadorXLSX:
    def __init__(self, agrupados):
        self.agrupados = agrupados

    def exportar(self):
        def calcular_segundos(entrada_str, salida_str):
            try:
                e = datetime.strptime(entrada_str.replace("📝", "").strip(), "%H:%M")
                s = datetime.strptime(salida_str.replace("📝", "").strip(), "%H:%M")
                dur = (s - e).total_seconds()
                return max(0, int(dur))
            except:
                return 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Fichajes"

        # Estilos
        encabezado_font = Font(bold=True, color="FFFFFF")
        encabezado_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        borde = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))
        alineado = Alignment(horizontal="center", vertical="center")
        bold = Font(bold=True)
        fill_dia = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_usuario = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        columnas = ["Usuario", "Fecha", "Hora de Entrada", "Hora de Salida", "Duración"]
        ws.append(columnas)
        for col in range(1, len(columnas) + 1):
            c = ws.cell(row=1, column=col)
            c.font = encabezado_font
            c.fill = encabezado_fill
            c.border = borde
            c.alignment = alineado

        resumen_por_usuario = {}
        motivos_por_usuario = {}

        for entry in self.agrupados:
            usuario = entry.get("usuario", "")
            fecha = entry.get("fecha", "")
            intervalos = entry.get("intervalos", [])

            total_segundos_dia = 0

            for t in intervalos:
                entrada_str = t.get("entrada", "")
                salida_str = t.get("salida", "")

                if t.get("manualEntrada"):
                    entrada_str = f"📝 {entrada_str}"
                    if t.get("motivoEntrada"):
                        motivos_por_usuario.setdefault(usuario, []).append(f"{fecha} – Entrada manual: “{t['motivoEntrada']}”")
                if t.get("manualSalida"):
                    salida_str = f"📝 {salida_str}"
                    if t.get("motivoSalida"):
                        motivos_por_usuario.setdefault(usuario, []).append(f"{fecha} – Salida manual: “{t['motivoSalida']}”")

                duracion_segundos = calcular_segundos(entrada_str, salida_str)
                total_segundos_dia += duracion_segundos
                duracion_str = f"{duracion_segundos // 3600}h {int((duracion_segundos % 3600) / 60)}min"

                fila = [usuario, fecha, entrada_str, salida_str, duracion_str]
                ws.append(fila)

            if intervalos:
                h_dia = total_segundos_dia // 3600
                m_dia = (total_segundos_dia % 3600) // 60
                total_dia_str = f"{h_dia}h {m_dia}min" if h_dia else f"{m_dia}min"

                ws.append(["", "", "TOTAL DÍA", "", total_dia_str])
                last_row = ws.max_row
                for col in range(1, len(columnas) + 1):
                    c = ws.cell(row=last_row, column=col)
                    c.font = bold
                    c.fill = fill_dia
                    c.alignment = alineado
                    c.border = borde
                ws.append([""] * len(columnas))

                resumen_por_usuario[usuario] = resumen_por_usuario.get(usuario, 0) + total_segundos_dia

        for usuario, total_seg in resumen_por_usuario.items():
            h = total_seg // 3600
            m = (total_seg % 3600) // 60
            total_str = f"{h}h {m}min" if h else f"{m}min"
            ws.append([""] * len(columnas))
            ws.append(["", "", f"TOTAL USUARIO {usuario}", "", total_str])
            last_row = ws.max_row
            for col in range(1, len(columnas) + 1):
                c = ws.cell(row=last_row, column=col)
                c.font = bold
                c.fill = fill_usuario
                c.alignment = alineado
                c.border = borde

        # Bloque final de motivos manuales
        if motivos_por_usuario:
            ws.append([""] * len(columnas))
            ws.append(["📝 MOTIVOS DE FICHAJES MANUALES"])
            for usuario in sorted(motivos_por_usuario):
                ws.append([f"👤 {usuario}"])
                for motivo in motivos_por_usuario[usuario]:
                    ws.append([f"• {motivo}"])

        # Pie de página
        ws.append([""] * len(columnas))
        ws.append(["Documento generado automáticamente - No modificar manualmente"])
        ws.append([f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}"])
        for i in range(ws.max_row - 1, ws.max_row + 1):
            for j in range(1, 3):
                c = ws.cell(row=i, column=j)
                c.font = Font(italic=True, color="666666")
                c.alignment = Alignment(horizontal="left")

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=logs.xlsx"}
        )
